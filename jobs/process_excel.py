# Excel Ä°ÅŸleme GÃ¶revi (jobs/process_excel.py)

import asyncio
from pathlib import Path
from typing import Dict, Any, List
from openpyxl import load_workbook
import tempfile
import zipfile
from datetime import datetime

from utils.excel_cleaner import clean_excel_headers
from utils.excel_splitter import split_excel_by_groups
from utils.mailer import send_email_with_attachment
from utils.group_manager import group_manager
from utils.logger import logger
import config  # config modÃ¼lÃ¼nÃ¼ import etmeyi unutmayÄ±n

async def zip_and_send_output_folder(task_result: Dict, original_filename: str) -> bool:
    """Output klasÃ¶rÃ¼nÃ¼ zipleyip PERSONAL_EMAIL'e gÃ¶nderir"""
    try:
        if not config.config.PERSONAL_EMAIL:
            logger.warning("PERSONAL_EMAIL tanÄ±mlÄ± deÄŸil, zip gÃ¶nderilmeyecek")
            return False
        
        # Output klasÃ¶rÃ¼ kontrolÃ¼
        if not config.config.OUTPUT_DIR.exists() or not any(config.config.OUTPUT_DIR.iterdir()):
            logger.warning("Output klasÃ¶rÃ¼ boÅŸ, zip gÃ¶nderilmeyecek")
            return False
        
        # Zip dosyasÄ± oluÅŸtur
        timestamp = datetime.now().strftime("%m%d_%H%M")  # output_0924_2136 formatÄ±
        zip_filename = f"output_{timestamp}.zip"
        zip_path = config.config.OUTPUT_DIR.parent / zip_filename  # data klasÃ¶rÃ¼ne kaydet
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_path in config.config.OUTPUT_DIR.glob('*.xlsx'):
                zipf.write(file_path, file_path.name)
        
        # Mail gÃ¶nder
        subject = f"ğŸ“Š Output KlasÃ¶rÃ¼ - {original_filename}"
        body = (
            f"Merhaba,\n\n"
            f"{original_filename} iÅŸlemi tamamlandÄ±.\n"
            f"Output klasÃ¶rÃ¼ndeki {len(list(config.config.OUTPUT_DIR.glob('*.xlsx')))} dosya ektedir.\n\n"
            f"Ä°ÅŸlem detaylarÄ±:\n"
            f"- Toplam satÄ±r: {task_result.get('total_rows', 0)}\n"
            f"- EÅŸleÅŸen satÄ±r: {task_result.get('matched_rows', 0)}\n"
            f"- OluÅŸan grup: {len(task_result.get('output_files', {}))}\n\n"
            f"Ä°yi Ã§alÄ±ÅŸmalar,\nExcel Bot"
        )
        
        success = await send_email_with_attachment(
            [config.config.PERSONAL_EMAIL], subject, body, zip_path
        )
        
        # GeÃ§ici zip dosyasÄ±nÄ± sil
        zip_path.unlink()
        
        if success:
            logger.info(f"Output zip baÅŸarÄ±yla gÃ¶nderildi: {zip_filename}")
        else:
            logger.error(f"Output zip gÃ¶nderilemedi: {zip_filename}")
            
        return success
        
    except Exception as e:
        logger.error(f"Output zip gÃ¶nderme hatasÄ±: {e}")
        return False

async def process_excel_task(input_path: Path, user_id: int) -> Dict[str, Any]:
    """Excel iÅŸleme gÃ¶revini yÃ¼rÃ¼tÃ¼r (output zip Ã¶zellikli)"""
    cleaning_result = None
    original_filename = input_path.name  # Orijinal dosya adÄ±nÄ± kaydet
    
    try:
        logger.info(f"Excel iÅŸleme baÅŸlatÄ±ldÄ±: {input_path.name}, KullanÄ±cÄ±: {user_id}")

        # 1. Excel dosyasÄ±nÄ± temizle ve dÃ¼zenle
        cleaning_result = clean_excel_headers(str(input_path))
        if not cleaning_result["success"]:
            error_msg = f"Excel temizleme hatasÄ±: {cleaning_result.get('error', 'Bilinmeyen hata')}"
            logger.error(error_msg)
            return {"success": False, "error": error_msg}
        
        logger.info(f"Excel temizlendi: {cleaning_result['row_count']} satÄ±r")

        # 2. DosyayÄ± gruplara ayÄ±r
        splitting_result = split_excel_by_groups(
            cleaning_result["temp_path"],
            cleaning_result["headers"]
        )
        
        if not splitting_result["success"]:
            error_msg = f"Excel ayÄ±rma hatasÄ±: {splitting_result.get('error', 'Bilinmeyen hata')}"
            logger.error(error_msg)
            return {"success": False, "error": error_msg}
        
        logger.info(f"Excel gruplara ayrÄ±ldÄ±: {splitting_result['total_rows']} satÄ±r, {len(splitting_result['output_files'])} grup")

        # 3. E-postalarÄ± gÃ¶nder (async olarak)
        email_tasks = []
        output_files = splitting_result["output_files"]
        email_results = []
        
        for group_id, file_info in output_files.items():
            group_info = group_manager.get_group_info(group_id)
            recipients = group_info.get("email_recipients", [])
            
            if recipients and file_info["row_count"] > 0:
                subject = f"{group_info.get('group_name', group_id)} Raporu - {file_info['filename']}"
                body = (
                    f"Merhaba,\n\n"
                    f"{group_info.get('group_name', group_id)} grubu iÃ§in {file_info['row_count']} satÄ±rlÄ±k rapor ekte gÃ¶nderilmiÅŸtir.\n\n"
                    f"Ä°yi Ã§alÄ±ÅŸmalar,\nExcel Bot"
                )
                
                # Her alÄ±cÄ± iÃ§in ayrÄ± mail gÃ¶nderimi
                for recipient in recipients:
                    if recipient.strip():  # BoÅŸ email adreslerini atla
                        task = send_email_with_attachment(
                            [recipient.strip()], subject, body, file_info["path"]
                        )
                        email_tasks.append((task, group_id, recipient, file_info["path"].name))
        
        # TÃ¼m mail gÃ¶revlerini paralel Ã§alÄ±ÅŸtÄ±r
        if email_tasks:
            logger.info(f"{len(email_tasks)} mail gÃ¶revi baÅŸlatÄ±lÄ±yor...")
            
            # GÃ¶revleri topla ve Ã§alÄ±ÅŸtÄ±r
            tasks = [task[0] for task in email_tasks]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            # SonuÃ§larÄ± iÅŸle
            for i, result in enumerate(results):
                task_info = email_tasks[i]
                group_id, recipient, filename = task_info[1], task_info[2], task_info[3]
                
                if isinstance(result, Exception):
                    logger.error(f"Mail gÃ¶nderim hatasÄ± - Grup: {group_id}, AlÄ±cÄ±: {recipient}, Dosya: {filename}, Hata: {result}")
                    email_results.append({
                        "success": False,
                        "group_id": group_id,
                        "recipient": recipient,
                        "error": str(result)
                    })
                else:
                    logger.info(f"Mail gÃ¶nderildi - Grup: {group_id}, AlÄ±cÄ±: {recipient}, Dosya: {filename}")
                    email_results.append({
                        "success": True,
                        "group_id": group_id,
                        "recipient": recipient
                    })
            
            # Grup bazÄ±nda email durumunu gÃ¼ncelle
            successful_emails = sum(1 for res in email_results if res["success"])
            logger.info(f"Mail gÃ¶nderim sonucu: {successful_emails} baÅŸarÄ±lÄ±, {len(email_results) - successful_emails} baÅŸarÄ±sÄ±z")
        
        # 4. GeÃ§ici dosyalarÄ± temizle
        try:
            if cleaning_result and "temp_path" in cleaning_result:
                temp_path = Path(cleaning_result["temp_path"])
                if temp_path.exists():
                    temp_path.unlink()
                    logger.info(f"GeÃ§ici dosya silindi: {temp_path.name}")
        except Exception as e:
            logger.warning(f"GeÃ§ici dosya silinemedi: {e}")
        
        # BaÅŸarÄ±lÄ± task sonucunu oluÅŸtur
        task_result = {
            "success": True,
            "output_files": output_files,
            "total_rows": splitting_result["total_rows"],
            "matched_rows": splitting_result["matched_rows"],
            "email_results": email_results,
            "user_id": user_id
        }
        
        # 5. Output klasÃ¶rÃ¼nÃ¼ zipleyip gÃ¶nder (YENÄ° EKLENEN KISIM)
        if task_result["success"]:
            try:
                await zip_and_send_output_folder(task_result, original_filename)
            except Exception as e:
                logger.error(f"Output zip gÃ¶nderim hatasÄ±: {e}")
                # Zip hatasÄ± ana iÅŸlemi baÅŸarÄ±sÄ±z yapmasÄ±n
        
        return task_result
        
    except Exception as e:
        logger.error(f"Ä°ÅŸlem gÃ¶revi hatasÄ±: {e}", exc_info=True)
        
        # Hata durumunda geÃ§ici dosyalarÄ± temizle
        try:
            if cleaning_result and "temp_path" in cleaning_result:
                temp_path = Path(cleaning_result["temp_path"])
                if temp_path.exists():
                    temp_path.unlink()
        except:
            pass
            
        return {"success": False, "error": str(e)}

#kiÅŸisel mail fonksiyonu
async def process_excel_task_for_personal_email(input_path: Path, user_id: int) -> Dict[str, Any]:
    """Sadece kiÅŸisel maile gÃ¶nderim iÃ§in Excel iÅŸleme gÃ¶revi"""
    cleaning_result = None
    try:
        logger.info(f"KiÅŸisel mail iÃ§in Excel iÅŸleme baÅŸlatÄ±ldÄ±: {input_path.name}, KullanÄ±cÄ±: {user_id}")

        # 1. Excel dosyasÄ±nÄ± temizle ve dÃ¼zenle (aynÄ± iÅŸlem)
        cleaning_result = clean_excel_headers(str(input_path))
        if not cleaning_result["success"]:
            error_msg = f"Excel temizleme hatasÄ±: {cleaning_result.get('error', 'Bilinmeyen hata')}"
            logger.error(error_msg)
            return {"success": False, "error": error_msg}
        
        logger.info(f"Excel temizlendi: {cleaning_result['row_count']} satÄ±r")

        # 2. TÃ¼m verileri tek bir dosyada tut (gruplara ayÄ±rma YOK)
        # TemizlenmiÅŸ dosyayÄ± yÃ¼kle
        wb = load_workbook(cleaning_result["temp_path"])
        ws = wb.active
        
        # SÃ¼tun geniÅŸliklerini ayarla
        from openpyxl.utils import get_column_letter
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            column_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[column_letter].width = min(25, max(length + 2, 10))
        
        # GeÃ§ici Ã§Ä±ktÄ± dosyasÄ± oluÅŸtur
        temp_output = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_output_path = temp_output.name
        wb.save(temp_output_path)
        wb.close()
        
        # 3. Sadece kiÅŸisel maile gÃ¶nder
        email_success = False
        if config.config.PERSONAL_EMAIL:
            subject = f"ğŸ“Š Excel Raporu - {input_path.name}"
            body = (
                f"Merhaba,\n\n"
                f"{cleaning_result['row_count']} satÄ±rlÄ±k Excel raporu ekte gÃ¶nderilmiÅŸtir.\n\n"
                f"Ä°yi Ã§alÄ±ÅŸmalar,\nExcel Bot"
            )
            
            email_success = await send_email_with_attachment(
                [config.config.PERSONAL_EMAIL], subject, body, Path(temp_output_path)
            )
        
        # 4. GeÃ§ici dosyalarÄ± temizle
        try:
            if cleaning_result and "temp_path" in cleaning_result:
                temp_path = Path(cleaning_result["temp_path"])
                if temp_path.exists():
                    temp_path.unlink()
            if Path(temp_output_path).exists():
                Path(temp_output_path).unlink()
        except Exception as e:
            logger.warning(f"GeÃ§ici dosya silinemedi: {e}")
        
        return {
            "success": email_success,
            "total_rows": cleaning_result["row_count"],
            "email_sent_to": config.config.PERSONAL_EMAIL if email_success else None,
            "user_id": user_id
        }
        
    except Exception as e:
        logger.error(f"KiÅŸisel mail iÅŸleme hatasÄ±: {e}", exc_info=True)
        
        # Hata durumunda geÃ§ici dosyalarÄ± temizle
        try:
            if cleaning_result and "temp_path" in cleaning_result:
                temp_path = Path(cleaning_result["temp_path"])
                if temp_path.exists():
                    temp_path.unlink()
            if 'temp_output_path' in locals() and Path(temp_output_path).exists():
                Path(temp_output_path).unlink()
        except:
            pass
            
        return {"success": False, "error": str(e)}