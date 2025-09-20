#Excel Ayırıcı (utils/excel_splitter.py)
from openpyxl import load_workbook, Workbook
from typing import Dict, List, Tuple, Any

from utils.group_manager import group_manager
from utils.file_namer import generate_output_filename
from utils.logger import logger
from config import config
import os

class ExcelSplitter:
    def __init__(self):
        self.workbooks = {}  # group_id -> Workbook
        self.sheets = {}     # group_id -> Worksheet
        self.row_counts = {} # group_id -> satır sayısı
        self.headers = []    # başlık satırı
    
    def initialize_workbook(self, group_id: str):
        """Yeni bir workbook ve worksheet oluşturur"""
        if group_id not in self.workbooks:
            wb = Workbook()
            ws = wb.active
            ws.title = "Veriler"
            
            # Başlık satırını yaz
            for col_idx, header in enumerate(self.headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            self.workbooks[group_id] = wb
            self.sheets[group_id] = ws
            self.row_counts[group_id] = 1  # Başlık satırı
    
    def process_excel_file(self, input_path: str, headers: List[str]) -> Dict[str, Any]:
        """Excel dosyasını gruplara ayırır"""
        try:
            self.headers = headers
            
            # Sadece okuma modunda aç
            wb = load_workbook(filename=input_path, read_only=True)
            ws = wb.active
            
            # Satırları işle (chunk bazlı)
            chunk_size = 1000
            processed_rows = 0
            
            for row_idx in range(2, ws.max_row + 1):  # Başlık satırını atla
                # Satırı oku
                row_data = []
                for col in range(1, len(headers) + 1):
                    cell_value = ws.cell(row=row_idx, column=col).value
                    row_data.append(cell_value)
                
                # İl bilgisini al (B sütunu - index 1)
                city = row_data[1] if len(row_data) > 1 else None
                
                # Grubu belirle
                group_id = group_manager.get_group_for_city(city)
                
                # Workbook'u hazırla
                self.initialize_workbook(group_id)
                
                # Satırı ilgili gruba yaz
                ws_dest = self.sheets[group_id]
                current_row = self.row_counts[group_id] + 1
                
                for col_idx, value in enumerate(row_data, 1):
                    ws_dest.cell(row=current_row, column=col_idx, value=value)
                
                self.row_counts[group_id] = current_row
                processed_rows += 1
                
                # Chunk işleme - bellek yönetimi
                if processed_rows % chunk_size == 0:
                    logger.info(f"{processed_rows} satır işlendi")
            
            # Dosyaları kaydet
            output_files = {}
            for group_id, wb in self.workbooks.items():
                group_info = group_manager.get_group_info(group_id)
                filename = generate_output_filename(group_info)
                filepath = config.OUTPUT_DIR / filename
                
                wb.save(filepath)
                output_files[group_id] = {
                    "path": filepath,
                    "row_count": self.row_counts[group_id] - 1,  # Başlık hariç
                    "filename": filename
                }
            
            return {
                "success": True,
                "output_files": output_files,
                "total_rows": processed_rows,
                "matched_rows": processed_rows - self.row_counts.get("Grup_0", 1) + 1
            }
            
        except Exception as e:
            logger.error(f"Excel ayırma hatası: {e}")
            return {"success": False, "error": str(e)}
        finally:
            # Belleği temizle
            if 'wb' in locals():
                wb.close()
            self.close_all_workbooks()
    
    def close_all_workbooks(self):
        """Tüm workbook'ları kapatır"""
        for wb in self.workbooks.values():
            wb.close()
        self.workbooks.clear()
        self.sheets.clear()
        self.row_counts.clear()

def split_excel_by_groups(input_path: str, headers: List[str]) -> Dict[str, Any]:
    """Excel dosyasını gruplara ayıran ana fonksiyon"""
    splitter = ExcelSplitter()
    return splitter.process_excel_file(input_path, headers)
